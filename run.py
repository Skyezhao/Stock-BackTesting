import os
from rqalpha.api import *
from rqalpha import run_func
import xlrd
import numpy as np
import xlsxwriter
from openpyxl import load_workbook
from datetime import datetime
from xlrd import xldate_as_tuple
import openpyxl
import csv
import pandas as pd
from pandas import Series, DataFrame
zhibiao_list = ["ABS在建工程增幅","ABS在建工程增幅比ABS经营现金流增幅","ABS在建工程增幅比经营现金流增幅","ABS在建工程差值比固定资产原值","在建工程比固定资产原值","ABS在建工程差值比净利润"]
nn = 5
#0,3,5小
zhibiao = zhibiao_list[nn]
net_file = "results/"+ zhibiao +".xlsx"
data_file = "data/"+zhibiao+".xlsx"
construct = pd.read_excel('code/progress.xlsx', sheet_name = 0, index_col="Date")
pcf = pd.read_excel('code/progress.xlsx', sheet_name = 1,index_col="Date")
const_asset = pd.read_excel('code/progress.xlsx', sheet_name = 2,index_col="Date")
# tax_asset = pd.read_excel('code/progress.xlsx', sheet_name = 3,index_col="Date")
# opt_project = pd.read_excel('code/progress.xlsx', sheet_name = 4,index_col="Date")
profit = pd.read_excel('code/progress.xlsx', sheet_name = 5,index_col="Date")
# noncurrent = pd.read_excel('code/progress.xlsx', sheet_name = 6,index_col="Date")
# debt = pd.read_excel('code/progress.xlsx', sheet_name = 7,index_col="Date")

def create_excel(filename):
    excel = pd.DataFrame()
    excel.to_excel(filename)

def write_excel(filename,output):
    df = pd.DataFrame(output)
    writer=pd.ExcelWriter(filename)
    df.to_excel(writer, sheet_name="sheet1",header=False,index=False)
    writer.save()

#打开excel并写入
def write_in_excel(output_list, filename, line):
    wb = load_workbook(filename)
    wb1 = wb.active
    for i in range(len(output_list)):
        wb1.cell(line,i+1,output_list[i])
    wb.save(filename)

def re_stock(stock):
    stock = stock[:-3]
    if stock[0] == '6':
        stock += ".XSHG"
    else:
        stock += ".XSHE"
    return stock

stock_list = list(construct.index)

def calculate_zengfu(df):
    date_list = list(df.head())
    date_list = [time.strftime("%Y-%m-%d") for time in date_list]
    #-------- 计算数据 -----------
    output = []
    output.append(["Date"]+date_list[1:])
    for stock in stock_list:
        data = list(df.loc[stock])
        tmp = [stock]
        for k in range(1,len(data)):
            if (not np.isnan(data[k-1])) and int(data[k-1]) != 0:
                tmp.append(float((float(data[k])-float(data[k-1]))/float(data[k-1])))
            else:
                tmp.append("")
        output.append(tmp)
    return output

def ABS(list_):
    output = []
    output.append(list_[0])
    for i in range(1,len(list_)):
        tmp = [list_[i][0]]
        for k in range(1,len(list_[i])):
            if list_[i][k] != '':
                tmp.append(abs(float(list_[i][k])))
            else:
                tmp.append("")
        output.append(tmp)
    return output

def delete_one(df):
    date_list = list(df.head())
    date_list = [time.strftime("%Y-%m-%d") for time in date_list]
    #-------- 计算数据 -----------
    output = []
    output.append(["Date"]+date_list[1:])
    for stock in stock_list:
        data = list(df.loc[stock])[1:]
        tmp = [stock]
        for value in data:
            if np.isnan(value):
                tmp.append("")
            else:
                tmp.append(value)
        output.append(tmp)
    return output

def delete_none(df):
    date_list = list(df.head())
    date_list = [time.strftime("%Y-%m-%d") for time in date_list]
    #-------- 计算数据 -----------
    output = []
    output.append(["Date"]+date_list)
    for stock in stock_list:
        data = list(df.loc[stock])
        tmp = [stock]
        for value in data:
            if np.isnan(value):
                tmp.append("")
            else:
                tmp.append(value)
        output.append(tmp)
    return output

def calculate_chazhi(df):
    date_list = list(df.head())
    date_list = [time.strftime("%Y-%m-%d") for time in date_list]
    #-------- 计算数据 -----------
    output = []
    output.append(["Date"]+date_list[1:])
    for stock in stock_list:
        data = list(df.loc[stock])
        tmp = [stock]
        for k in range(1,len(data)):
            if not np.isnan(data[k-1]):
                tmp.append(float(float(data[k])-float(data[k-1])))
            else:
                tmp.append("")
        output.append(tmp)
    return output


def divide(list1, list2):
    output = []
    output.append(list1[0])
    for i in range(1,len(list1)):
        tmp = [list1[i][0]]
        for k in range(1,len(list1[i])):
            # print(list1[i][k],list2[i][k])
            if list2[i][k] != '' and list1[i][k] != '' and int(list2[i][k]) != 0:
                tmp.append(float(float(list1[i][k])/float(list2[i][k])))
            else:
                tmp.append("")
        output.append(tmp)
    return output


def get_data(number):
    if number == 0:
        #(本期在建工程-上期在建工程)/上期在建工程
        #-------- 读取csv文件 -----------
        # file.index = pd.to_datetime(file.index)
        create_excel(data_file)
        df1 = pd.DataFrame(construct)
        output = ABS(calculate_zengfu(df1))
        write_excel(data_file,output)
        #-------- 存储数值 -----------
        dic = {}
        year_list = []
        for i in range(len(output[0])):
            day = output[0][i][:4]
            if day == "Date":
                continue
            year_list.append(day)
            dic[day] = []
            for k in range(len(stock_list)):
                if output[k+1][i] != '':
                    dic[day].append((re_stock(stock_list[k]),(output[k+1][i])))
        return dic, year_list
    elif number == 1:
        #在建工程增幅/经营现金流增幅
        create_excel(data_file)
        df1 = pd.DataFrame(construct)
        construct2 = ABS(calculate_zengfu(df1))
        df2 = pd.DataFrame(pcf)
        pcf2 = ABS(calculate_zengfu(df2))
        output = divide(construct2, pcf2)
        write_excel(data_file,output)
        #-------- 存储数值 -----------
        dic = {}
        year_list = []
        for i in range(len(output[0])):
            day = output[0][i][:4]
            if day == "Date":
                continue
            year_list.append(day)
            dic[day] = []
            for k in range(len(stock_list)):
                if output[k+1][i] != '':
                    dic[day].append((re_stock(stock_list[k]),(output[k+1][i])))
        return dic, year_list
    elif number == 2:
        #在建工程增幅/经营现金流增幅 绝对值
        create_excel(data_file)
        df1 = pd.DataFrame(construct)
        construct2 = calculate_zengfu(df1)
        df2 = pd.DataFrame(pcf)
        pcf2 = calculate_zengfu(df2)
        output = ABS(divide(construct2, pcf2))
        write_excel(data_file,output)
        #-------- 存储数值 -----------
        dic = {}
        year_list = []
        for i in range(len(output[0])):
            day = output[0][i][:4]
            if day == "Date":
                continue
            year_list.append(day)
            dic[day] = []
            for k in range(len(stock_list)):
                if output[k+1][i] != '':
                    dic[day].append((re_stock(stock_list[k]),(output[k+1][i])))
        return dic, year_list
    elif number == 3:
        #ABS(本期在建工程-上一期在建工程)/固定资产原值
        create_excel(data_file)
        df1 = pd.DataFrame(construct)
        construct2 = calculate_chazhi(df1)
        df2 = pd.DataFrame(const_asset)
        const_asset2 = delete_one(df2)
        output = ABS(divide(construct2, const_asset2))
        write_excel(data_file,output)
        #-------- 存储数值 -----------
        dic = {}
        year_list = []
        for i in range(len(output[0])):
            day = output[0][i][:4]
            if day == "Date":
                continue
            year_list.append(day)
            dic[day] = []
            for k in range(len(stock_list)):
                if output[k+1][i] != '':
                    dic[day].append((re_stock(stock_list[k]),(output[k+1][i])))
        return dic, year_list
    elif number == 4:
        #在建工程/固定资产原值
        create_excel(data_file)
        df1 = pd.DataFrame(construct)
        construct2 = delete_none(df1)
        df2 = pd.DataFrame(const_asset)
        const_asset2 = delete_none(df2)
        output = divide(construct2, const_asset2)
        write_excel(data_file,output)
        #-------- 存储数值 -----------
        dic = {}
        year_list = []
        for i in range(len(output[0])):
            day = output[0][i][:4]
            if day == "Date":
                continue
            year_list.append(day)
            dic[day] = []
            for k in range(len(stock_list)):
                if output[k+1][i] != '':
                    dic[day].append((re_stock(stock_list[k]),(output[k+1][i])))
        return dic, year_list
    elif number == 5:
        #ABS(本期在建工程-上一期在建工程)/净利润
        create_excel(data_file)
        df1 = pd.DataFrame(construct)
        construct2 = calculate_chazhi(df1)
        df2 = pd.DataFrame(profit)
        profit2 = delete_one(df2)
        output = ABS(divide(construct2, profit2))
        write_excel(data_file,output)
        #-------- 存储数值 -----------
        dic = {}
        year_list = []
        for i in range(len(output[0])):
            day = output[0][i][:4]
            if day == "Date":
                continue
            year_list.append(day)
            dic[day] = []
            for k in range(len(stock_list)):
                if output[k+1][i] != '':
                    dic[day].append((re_stock(stock_list[k]),(output[k+1][i])))
        return dic, year_list

def divide_group(context,year):
    group = [[],[],[],[],[]]
    dic = context.dic #总： year-（stock，size）
    size_value = [] #存放size总列表
    #读取对应年份的全部（stock，size）元组
    #股票池内的股票才加入size_value列表中
    for pair in dic[year]:
        if pair[0] in context.stock:
            size_value.append(float(pair[1]))
    sort_values = np.sort(size_value)
    #计算四个分位点
    pin1 = float(sort_values[int((len(sort_values)+1)*0.2)])
    pin2 = float(sort_values[int((len(sort_values)+1)*0.4)])
    pin3 = float(sort_values[int((len(sort_values)+1)*0.6)])
    pin4 = float(sort_values[int((len(sort_values)+1)*0.8)])
    #分组
    for pair in dic[year]:
        if pair[1] < pin1:
            group[0].append(pair[0])
        elif pair[1] >= pin1 and pair[1] < pin2:
            group[1].append(pair[0])
        elif pair[1] >= pin2 and pair[1] < pin3:
            group[2].append(pair[0])
        elif pair[1] >= pin3 and pair[1] < pin4:
            group[3].append(pair[0])
        elif pair[1] >= pin4 :
            group[4].append(pair[0])
    print(len(group[0]),len(group[1]),len(group[2]),len(group[3]),len(group[4]))
    return group

def write_csv(output, filename):
    df = pd.DataFrame(output)
    df.to_csv(filename,index=False,header=False)

# 在这个方法中编写任何的初始化逻辑。context对象将会在你的算法策略的任何方法之间做传递。
def init(context):
    logger.info("init")
    context.net_list = [] #存储每日净值数据
    context.day_list = [] #存储时间轴
    context.lastbuy = [] #存储需要下一次需要被平仓的股票组合
    context.stock = [re_stock(x) for x in stock_list]
    context.dic, context.year_list = get_data(nn)
    for i in range(len(context.stock)):
        update_universe(context.stock[i])


def before_trading(context):
    pass


# 你选择的证券的数据更新将会触发此段逻辑，例如日或分钟历史数据切片或者是实时数据切片更新
def handle_bar(context, bar_dict):
    date = context.now.strftime('%Y-%m-%d')
    pre_date = get_previous_trading_date(date).strftime('%Y-%m-%d')
    year = str(int(date[:4])-1)
    # 定位每一年的4月底的下一个交易日，即五月第一天
    if date[5:7] == '05' and pre_date[5:7] == '04' and (year in context.year_list):
        group = divide_group(context,year)
        group_copy = group[context.turn-1]
        # print(group_copy)
        #----------平仓并提取需加仓股票------------
        #需要卖掉不在group_copy里的所有股票
        if context.lastbuy != []:
            last = np.array(context.lastbuy) #读取上次购买股票
            now = np.array(group_copy) #读取现在需持有股票
            sold_list = np.setdiff1d(last,now) #在上次仓位里但不在预期仓位里的股票，需要平仓
            for stock in sold_list:
                order_target_value(stock,0)
                # print("平仓： ",stock," 剩余金额： ",context.portfolio.cash)
        # #------------加仓---------------
        context.lastbuy = []
        shares = (context.portfolio.market_value+context.portfolio.cash)/(len(group_copy)*2) #计算每个股票的平均投入资产
        for stock in group_copy:
            result = order_target_value(stock,shares)
            if result != None:
                context.lastbuy.append(stock)
                # print("买入",stock," 现在现金： ",context.portfolio.cash)
    #存储日期及每日单位净值
    context.day_list.append(date)
    context.net_list.append(context.portfolio.unit_net_value)

def after_trading(context):
    next_date = get_next_trading_date(context.now.strftime('%Y-%m-%d')).strftime('%Y-%m-%d')
    if next_date > context.end:
        if context.turn == 1:
            write_in_excel(context.day_list,net_file ,1)
            write_in_excel(context.net_list,net_file ,2)
        else:
            write_in_excel(context.net_list,net_file ,3)